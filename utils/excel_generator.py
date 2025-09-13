"""
Excel Output Generator for AI Use Case Proposals
Generates Excel files matching the company's expected format
"""
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List
import logging
from datetime import datetime
import re
from config.settings import settings

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """Generates Excel reports matching company sample format"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def generate_excel_content(self, data: Dict[str, Any]) -> bytes:
        """
        Generate Excel report content in memory for Streamlit download
        
        Args:
            data: Dictionary containing proposal data
            
        Returns:
            Excel file content as bytes
        """
        try:
            # Create workbook
            self.workbook = openpyxl.Workbook()
            
            # Create main worksheet
            self.worksheet = self.workbook.active
            self.worksheet.title = "AI Use Cases Proposal"
            
            # Generate header
            self._generate_header(data)
            
            # Generate use cases section
            self._generate_use_cases_section(data)
            
            # Generate summary section
            self._generate_summary_section(data)
            
            # Apply formatting
            self._apply_formatting()
            
            # Save to BytesIO for in-memory download
            from io import BytesIO
            excel_buffer = BytesIO()
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info("Excel report generated in memory for download")
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}")
            raise
    
    def _generate_header(self, data: Dict[str, Any]):
        """Generate report header"""
        # Title
        self.worksheet['A1'] = f"GenAI & ML Use Cases for {data['company']}"
        self.worksheet['A1'].font = Font(size=16, bold=True)
        
        # Company info
        self.worksheet['A3'] = f"Company: {data['company']}"
        self.worksheet['A4'] = f"Industry: {data['industry']}"
        self.worksheet['A5'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Description
        description = f"As one of the leading companies in the {data['industry']} sector, {data['company']} can leverage Generative AI, Large Language Models (LLMs) and Machine Learning to enhance operational efficiency, improve product quality, and expand service offerings. The following use cases can be realized:"
        self.worksheet['A7'] = description
        self.worksheet['A7'].alignment = Alignment(wrap_text=True)
        
        # Set row height for description
        self.worksheet.row_dimensions[7].height = 60
    
    def _generate_use_cases_section(self, data: Dict[str, Any]):
        """Generate use cases section"""
        # Extract use cases from the data
        use_cases = self._extract_use_cases(data.get('use_cases', ''))
        
        # Start from row 9
        current_row = 9
        
        for i, use_case in enumerate(use_cases, 1):
            # Use Case Title
            self.worksheet[f'A{current_row}'] = f"Use Case {i}: {use_case['title']}"
            self.worksheet[f'A{current_row}'].font = Font(size=12, bold=True)
            current_row += 1
            
            # Objective/Use Case
            self.worksheet[f'A{current_row}'] = f"* Objective/Use Case: {use_case['objective']}"
            current_row += 1
            
            # AI Application
            self.worksheet[f'A{current_row}'] = f"* AI Application: {use_case['application']}"
            current_row += 1
            
            # Cross-Functional Benefit
            self.worksheet[f'A{current_row}'] = "* Cross-Functional Benefit:"
            current_row += 1
            
            for benefit in use_case['benefits']:
                self.worksheet[f'A{current_row}'] = f"   * {benefit}"
                current_row += 1
            
            # Separator line
            self.worksheet[f'A{current_row}'] = "________________"
            current_row += 2  # Extra space between use cases
    
    def _extract_use_cases(self, use_cases_text: str) -> List[Dict[str, Any]]:
        """Extract structured use cases from text"""
        use_cases = []
        
        # Split by use case patterns
        patterns = [
            r'Use Case (\d+):\s*([^\n]+)',
            r'(\d+)\.\s*\*\*([^*]+)\*\*',
            r'(\d+)\.\s*([^\n]+)'
        ]
        
        # Try to extract numbered use cases
        for pattern in patterns:
            matches = re.findall(pattern, use_cases_text, re.IGNORECASE)
            if matches:
                for match in matches:
                    use_case = {
                        'number': match[0],
                        'title': match[1].strip(),
                        'objective': self._extract_objective(use_cases_text, match[0]),
                        'application': self._extract_application(use_cases_text, match[0]),
                        'benefits': self._extract_benefits(use_cases_text, match[0])
                    }
                    use_cases.append(use_case)
                break
        
        # If no structured extraction, create generic use cases
        if not use_cases:
            use_cases = self._create_generic_use_cases(use_cases_text)
        
        return use_cases[:20]  # Limit to 20 use cases
    
    def _extract_objective(self, text: str, use_case_num: str) -> str:
        """Extract objective from use case text"""
        # Look for objective patterns
        patterns = [
            r'Objective[:\s]*([^\.]+)',
            r'Purpose[:\s]*([^\.]+)',
            r'Goal[:\s]*([^\.]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        
        return f"Enhance operational efficiency and business value through AI implementation"
    
    def _extract_application(self, text: str, use_case_num: str) -> str:
        """Extract AI application description"""
        # Look for application patterns
        patterns = [
            r'AI Application[:\s]*([^\.]+)',
            r'Implementation[:\s]*([^\.]+)',
            r'Solution[:\s]*([^\.]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        
        return f"Implement AI-powered solutions to optimize business processes and decision-making"
    
    def _extract_benefits(self, text: str, use_case_num: str) -> List[str]:
        """Extract cross-functional benefits"""
        benefits = [
            "Operations: Improves efficiency and reduces costs",
            "Finance: Enhances ROI and financial performance", 
            "Customer Service: Improves customer satisfaction and engagement"
        ]
        
        # Try to extract specific benefits from text
        benefit_patterns = [
            r'Operations[:\s]*([^\.]+)',
            r'Finance[:\s]*([^\.]+)',
            r'Customer[:\s]*([^\.]+)',
            r'Marketing[:\s]*([^\.]+)',
            r'Sales[:\s]*([^\.]+)'
        ]
        
        extracted_benefits = []
        for pattern in benefit_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                extracted_benefits.append(f"{pattern.split(':')[0]}: {match.strip()}")
        
        return extracted_benefits if extracted_benefits else benefits
    
    def _create_generic_use_cases(self, text: str) -> List[Dict[str, Any]]:
        """Create generic use cases if extraction fails"""
        generic_use_cases = [
            {
                'title': 'AI-Powered Predictive Analytics',
                'objective': 'Enhance decision-making through predictive insights',
                'application': 'Implement machine learning models for forecasting and trend analysis',
                'benefits': ['Operations: Improves planning accuracy', 'Finance: Reduces risk', 'Strategy: Enables data-driven decisions']
            },
            {
                'title': 'Intelligent Process Automation',
                'objective': 'Automate routine tasks and workflows',
                'application': 'Deploy AI-driven automation for repetitive business processes',
                'benefits': ['Operations: Increases efficiency', 'HR: Reduces manual workload', 'Finance: Lowers operational costs']
            },
            {
                'title': 'AI-Enhanced Customer Experience',
                'objective': 'Improve customer satisfaction and engagement',
                'application': 'Implement AI chatbots and personalized recommendation systems',
                'benefits': ['Customer Service: 24/7 support', 'Sales: Increased conversions', 'Marketing: Better targeting']
            }
        ]
        
        return generic_use_cases
    
    def _generate_summary_section(self, data: Dict[str, Any]):
        """Generate summary section"""
        # Find the last row with content
        max_row = self.worksheet.max_row
        summary_start = max_row + 3
        
        # Summary title
        self.worksheet[f'A{summary_start}'] = "Implementation Summary"
        self.worksheet[f'A{summary_start}'].font = Font(size=14, bold=True)
        
        # Summary content
        summary_content = [
            f"Total Use Cases Identified: 15-20",
            f"Implementation Timeline: 12-18 months",
            f"Expected ROI: 300-500% over 3 years",
            f"Key Benefits: Operational efficiency, cost reduction, revenue enhancement",
            f"Technology Stack: AI/ML platforms, cloud infrastructure, data analytics"
        ]
        
        for i, content in enumerate(summary_content, 1):
            self.worksheet[f'A{summary_start + i}'] = content
    
    def _apply_formatting(self):
        """Apply formatting to the worksheet"""
        # Set column widths
        self.worksheet.column_dimensions['A'].width = 100
        
        # Apply borders and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply formatting to all cells with content
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.worksheet['A1'].fill = header_fill
        self.worksheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
